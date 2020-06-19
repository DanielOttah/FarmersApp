import sqlite3
from openpyxl import Workbook, load_workbook


wb = load_workbook('./farmersmarkets.xlsx')
wbFood = load_workbook('./foodDetails.xlsx')
connection = sqlite3.connect('data.db')
cursor = connection.cursor()

# Create farmer table
create_table = "CREATE TABLE IF NOT EXISTS farmers (id INTEGER PRIMARY KEY, name text,address text,contact text, lat int, long int)"
cursor.execute(create_table)
# Insert values into table
insert = "INSERT into farmers VALUES(NULL,?,?,?,?,?)"
for row in wb['farmersmarkets'].iter_rows(min_row=2, max_row=wb['farmersmarkets'].max_row, values_only=True):
    cursor.execute(insert, (row[3], row[6], row[12], row[15], row[16]))

# Create product table
create_table = "CREATE TABLE IF NOT EXISTS foodDetails (id INTEGER PRIMARY KEY,  name text,botname text,othernames text, imageurl text)"
cursor.execute(create_table)
# Insert values into table
insert = "INSERT into foodDetails VALUES(NULL,?,?,?,?)"
for row in wbFood['food'].iter_rows(min_row=2, max_row=wbFood['food'].max_row, values_only=True):
    cursor.execute(insert, (row[1], row[2], row[3], row[4]))


# def viewTable():
#     for row in cursor.execute("SELECT * FROM farmers"):
#         print("row")
#         print(row)


# viewTable()

connection.commit()
connection.close()
